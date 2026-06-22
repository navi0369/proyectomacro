#!/usr/bin/env python3
"""
Data extraction and Excel generation for BCB fiscal data 1981-1989.
Extracts: ingresos, egresos, deficit/superavit, deuda interna SPNF.
"""

import json
import re
import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages"], check=True)
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

BASE_DIR = "/home/navi/projects/proyectomacro"
OCR_DIR = f"{BASE_DIR}/data/ocr"
OUTPUT_DIR = f"{BASE_DIR}/output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

YEARS = list(range(1981, 1990))

# Currency info per year
# Bolivia used "Pesos Bolivianos" until Feb 1987 (Law 901, Dec 1986 introduced "Boliviano" = 1,000,000 pesos)
# Actually the "Peso Boliviano" was used 1963-1987, then "Boliviano" from 1987
# 1 Boliviano = 1,000,000 Pesos Bolivianos
CURRENCY_INFO = {
    1981: {"currency": "Pesos Bolivianos", "unit": "millones"},
    1982: {"currency": "Pesos Bolivianos", "unit": "millones"},
    1983: {"currency": "Pesos Bolivianos", "unit": "millones"},
    1984: {"currency": "Pesos Bolivianos", "unit": "millones"},
    1985: {"currency": "Pesos Bolivianos", "unit": "millones"},
    1986: {"currency": "Pesos Bolivianos", "unit": "millones"},
    1987: {"currency": "Bolivianos", "unit": "millones"},
    1988: {"currency": "Bolivianos", "unit": "millones"},
    1989: {"currency": "Bolivianos", "unit": "millones"},
}

def clean_number(text):
    """Try to parse a number from text, handling Spanish/Bolivian number formats."""
    if not text:
        return None
    # Remove spaces
    text = text.strip()
    # Handle negative in parentheses (accounting convention)
    negative = False
    if text.startswith('(') and text.endswith(')'):
        negative = True
        text = text[1:-1]
    if text.startswith('-'):
        negative = True
        text = text[1:]
    
    # Remove currency symbols and letters
    text = re.sub(r'[Bb]s\.?|Bs\.?|\$|[A-Za-z]', '', text).strip()
    
    # Handle different decimal separators
    # If format is like "1.234.567,89" -> European format
    # If format is like "1,234,567.89" -> US format
    # If format is like "1234567" -> plain
    
    if re.match(r'^\d{1,3}(\.\d{3})*,\d+$', text):
        # European: 1.234.567,89
        text = text.replace('.', '').replace(',', '.')
    elif re.match(r'^\d{1,3}(,\d{3})*\.\d+$', text):
        # US: 1,234,567.89
        text = text.replace(',', '')
    elif re.match(r'^\d{1,3}(\.\d{3})+$', text):
        # Thousands only: 1.234.567
        text = text.replace('.', '')
    elif re.match(r'^\d{1,3}(,\d{3})+$', text):
        # Thousands only: 1,234,567
        text = text.replace(',', '')
    else:
        # Remove all non-numeric except dot
        text = re.sub(r'[^\d\.]', '', text)
    
    try:
        val = float(text)
        return -val if negative else val
    except:
        return None


def find_fiscal_data_in_text(pages_data, year):
    """
    Search through OCR pages to find fiscal data tables.
    Returns dict with structured data and source references.
    """
    
    results = {
        "year": year,
        "currency": CURRENCY_INFO[year]["currency"],
        "unit": CURRENCY_INFO[year]["unit"],
        "ingresos_totales": None,
        "ingresos_corrientes": None,
        "ingresos_capital": None,
        "egresos_totales": None,
        "egresos_corrientes": None,
        "egresos_capital": None,
        "deficit_superavit": None,
        "deuda_interna_total": None,
        "deuda_interna_bcb": None,
        "deuda_interna_banca": None,
        "deuda_interna_otros": None,
        "sources": []
    }
    
    # Patterns to look for
    ingreso_patterns = [
        r'ingresos?\s+totales?\s*[:\|]?\s*([\d\.,\(\)\-]+)',
        r'total\s+ingresos?\s*[:\|]?\s*([\d\.,\(\)\-]+)',
        r'ingresos?\s+corrientes?\s*[:\|]?\s*([\d\.,\(\)\-]+)',
    ]
    egreso_patterns = [
        r'egresos?\s+totales?\s*[:\|]?\s*([\d\.,\(\)\-]+)',
        r'total\s+egresos?\s*[:\|]?\s*([\d\.,\(\)\-]+)',
        r'gastos?\s+totales?\s*[:\|]?\s*([\d\.,\(\)\-]+)',
    ]
    deficit_patterns = [
        r'd[eé]ficit\s*[:\|\/]?\s*super[aá]vit\s*[:\|]?\s*([\d\.,\(\)\-]+)',
        r'd[eé]ficit\s+(?:global|fiscal|total)\s*[:\|]?\s*([\d\.,\(\)\-]+)',
        r'resultado\s*[:\|]?\s*([\d\.,\(\)\-]+)',
        r'super[aá]vit\s*[:\|]?\s*([\d\.,\(\)\-]+)',
        r'd[eé]ficit\s*[:\|]?\s*([\d\.,\(\)\-]+)',
    ]
    deuda_patterns = [
        r'deuda\s+interna\s+(?:total|del\s+spnf|p[uú]blica)?\s*[:\|]?\s*([\d\.,\(\)\-]+)',
        r'cr[eé]dito\s+interno\s*(?:neto|total)?\s*[:\|]?\s*([\d\.,\(\)\-]+)',
        r'endeudamiento\s+interno\s*[:\|]?\s*([\d\.,\(\)\-]+)',
    ]
    
    all_relevant_text = []
    
    for page_data in pages_data:
        page_num = page_data["page"]
        text = page_data["text"]
        text_lower = text.lower()
        
        # Check relevance
        is_spnf = any(kw in text_lower for kw in [
            "sector p", "spnf", "publico no financiero", "público no financiero",
            "no financiero"
        ])
        is_fiscal = any(kw in text_lower for kw in [
            "ingresos", "egresos", "deficit", "déficit", "superavit",
            "resultado fiscal", "financiamiento"
        ])
        is_debt = any(kw in text_lower for kw in [
            "deuda interna", "endeudamiento interno", "credito interno"
        ])
        
        if not (is_fiscal or is_debt):
            continue
        
        # Record as source
        source_info = {
            "page": page_num,
            "year": year,
            "context": "SPNF fiscal" if is_spnf else ("debt" if is_debt else "fiscal"),
            "snippet": text[:500]
        }
        all_relevant_text.append((page_num, text, source_info))
        
        # Try to extract values
        lines = text.split('\n')
        for i, line in enumerate(lines):
            line_lower = line.lower()
            
            # Try ingreso patterns
            for pat in ingreso_patterns:
                m = re.search(pat, line_lower)
                if m:
                    val = clean_number(m.group(1))
                    if val and val > 0:
                        if 'corriente' in pat and results["ingresos_corrientes"] is None:
                            results["ingresos_corrientes"] = val
                            results["sources"].append(f"Ingresos corrientes: p.{page_num}")
                        elif results["ingresos_totales"] is None:
                            results["ingresos_totales"] = val
                            results["sources"].append(f"Ingresos totales: p.{page_num}")
            
            # Try egreso patterns
            for pat in egreso_patterns:
                m = re.search(pat, line_lower)
                if m:
                    val = clean_number(m.group(1))
                    if val and val > 0:
                        if results["egresos_totales"] is None:
                            results["egresos_totales"] = val
                            results["sources"].append(f"Egresos totales: p.{page_num}")
            
            # Try deficit patterns
            for pat in deficit_patterns:
                m = re.search(pat, line_lower)
                if m:
                    val = clean_number(m.group(1))
                    if val is not None and results["deficit_superavit"] is None:
                        results["deficit_superavit"] = val
                        results["sources"].append(f"Deficit/Superavit: p.{page_num}")
            
            # Try debt patterns
            for pat in deuda_patterns:
                m = re.search(pat, line_lower)
                if m:
                    val = clean_number(m.group(1))
                    if val and val > 0 and results["deuda_interna_total"] is None:
                        results["deuda_interna_total"] = val
                        results["sources"].append(f"Deuda interna: p.{page_num}")
    
    results["relevant_pages"] = [(p[0], p[2]) for p in all_relevant_text]
    return results, all_relevant_text


def create_excel(all_data, output_path):
    """Create formatted Excel with all fiscal data."""
    wb = openpyxl.Workbook()
    
    # =========================================================
    # Sheet 1: Resumen / Summary
    # =========================================================
    ws = wb.active
    ws.title = "Resumen SPNF 1981-1989"
    
    # Colors
    header_fill = PatternFill("solid", fgColor="1F4E79")
    subheader_fill = PatternFill("solid", fgColor="2E75B6")
    alt_fill = PatternFill("solid", fgColor="DEEAF1")
    
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    subheader_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    normal_font = Font(name="Calibri", size=10)
    title_font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = "SECTOR PÚBLICO NO FINANCIERO (SPNF) - DATOS FISCALES 1981-1989"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:L2")
    ws["A2"] = "Banco Central de Bolivia - Memorias Anuales | Fuente: BCB"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center")
    
    ws.merge_cells("A3:L3")
    ws["A3"] = "Nota: Pesos Bolivianos (Pb$) hasta 1986; Bolivianos (Bs.) desde 1987. 1 Boliviano = 1,000,000 Pesos Bolivianos"
    ws["A3"].font = Font(name="Calibri", italic=True, size=9, color="FF0000")
    ws["A3"].alignment = Alignment(horizontal="center")
    
    # Headers
    headers = ["Año", "Moneda", "Unidad", "Ingresos\nTotales", "Ingresos\nCorrientes",
               "Egresos\nTotales", "Egresos\nCorrientes", "Déficit (-)\nSuperávit (+)",
               "Deuda Int.\nTotal", "Deuda Int.\nBCB", "Deuda Int.\nBanca", "Fuentes"]
    
    col_widths = [8, 22, 12, 15, 15, 15, 15, 18, 15, 15, 15, 40]
    
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    
    ws.row_dimensions[5].height = 35
    
    # Data rows
    for row_idx, data in enumerate(all_data, 6):
        year = data["year"]
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        
        row_data = [
            year,
            data["currency"],
            data["unit"],
            data.get("ingresos_totales"),
            data.get("ingresos_corrientes"),
            data.get("egresos_totales"),
            data.get("egresos_corrientes"),
            data.get("deficit_superavit"),
            data.get("deuda_interna_total"),
            data.get("deuda_interna_bcb"),
            data.get("deuda_interna_banca"),
            "; ".join(data.get("sources", [])[:5])
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            cell.fill = fill
            
            if col_idx == 1:  # Year
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(name="Calibri", bold=True, size=10)
            elif isinstance(val, (int, float)) and val is not None and col_idx > 3:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
                if col_idx == 8 and val is not None:  # Deficit/Superavit
                    if val < 0:
                        cell.font = Font(name="Calibri", color="FF0000", size=10)
                    else:
                        cell.font = Font(name="Calibri", color="375623", size=10)
    
    # Freeze panes
    ws.freeze_panes = "A6"
    
    # =========================================================
    # Sheet 2: Pages Raw Data
    # =========================================================
    ws2 = wb.create_sheet("Páginas Relevantes")
    ws2["A1"] = "PÁGINAS RELEVANTES IDENTIFICADAS EN CADA MEMORIA BCB"
    ws2["A1"].font = title_font
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 80
    
    headers2 = ["Año", "Página", "Contexto", "Extracto de Texto"]
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    row2 = 4
    for data in all_data:
        year = data["year"]
        for page_num, source_info in data.get("relevant_pages", []):
            ws2.cell(row=row2, column=1, value=year).border = thin_border
            ws2.cell(row=row2, column=2, value=page_num).border = thin_border
            ws2.cell(row=row2, column=3, value=source_info.get("context", "")).border = thin_border
            snippet_cell = ws2.cell(row=row2, column=4, value=source_info.get("snippet", "")[:300])
            snippet_cell.border = thin_border
            snippet_cell.alignment = Alignment(wrap_text=True)
            ws2.row_dimensions[row2].height = 60
            row2 += 1
    
    # Save
    wb.save(output_path)
    print(f"Excel saved to: {output_path}")
    return output_path


def main():
    all_results = []
    
    for year in YEARS:
        json_path = f"{OCR_DIR}/bcb_{year}_ocr.json"
        if not os.path.exists(json_path):
            print(f"[{year}] No OCR data found at {json_path}")
            all_results.append({
                "year": year,
                "currency": CURRENCY_INFO[year]["currency"],
                "unit": CURRENCY_INFO[year]["unit"],
                "sources": [],
                "relevant_pages": []
            })
            continue
        
        print(f"\nAnalyzing {year}...")
        with open(json_path, 'r', encoding='utf-8') as f:
            pages_data = json.load(f)
        
        result, relevant_text = find_fiscal_data_in_text(pages_data, year)
        all_results.append(result)
        
        print(f"  Found {len(relevant_text)} relevant pages")
        print(f"  Ingresos: {result.get('ingresos_totales')}")
        print(f"  Egresos: {result.get('egresos_totales')}")
        print(f"  Deficit/Superavit: {result.get('deficit_superavit')}")
        print(f"  Deuda Interna: {result.get('deuda_interna_total')}")
    
    # Save intermediate results as JSON
    json_output = f"{OUTPUT_DIR}/bcb_fiscal_data.json"
    with open(json_output, 'w', encoding='utf-8') as f:
        # Remove large fields for JSON output
        clean_results = []
        for r in all_results:
            clean = {k: v for k, v in r.items() if k != 'relevant_pages'}
            clean_results.append(clean)
        json.dump(clean_results, f, ensure_ascii=False, indent=2)
    print(f"\nJSON data saved to: {json_output}")
    
    # Create Excel
    excel_output = f"{OUTPUT_DIR}/BCB_Fiscal_Data_1981_1989.xlsx"
    create_excel(all_results, excel_output)
    
    return all_results


if __name__ == "__main__":
    results = main()
